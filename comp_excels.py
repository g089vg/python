# -*- coding: utf-8 -*-
"""
Created on Tue Jan 28 18:46:30 2020

@author: mitan
"""
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string
from openpyxl.styles.fonts import Font
from openpyxl.styles.colors import Color



threshold = 5

#Excelファイルオープン
file_name1 = 'comp1.xlsx'
file_name2 = 'comp2.xlsx'

book1 = load_workbook(file_name1)
book2 = load_workbook(file_name2)
#シート名取得
sheet1 = book1.get_sheet_names()
sheet2 = book1.get_sheet_names()
#sheet = np.array(sheet)

dst_book = Workbook()
dst_book.remove_sheet(dst_book.get_sheet_by_name('Sheet'))

for sheet_number in range(len(sheet1)):
    sheet_set1 = book1[sheet1[sheet_number]]
    sheet_set2 = book2[sheet2[sheet_number]]


    excel_data_sheet1 = []
    excel_data_sheet2 = []
    # セル番地を取得 
    for cells in tuple(sheet_set1.rows):
        excel_data_cell1 = []
        for cell in cells:
            excel_data_cell1.append(cell.value)
        excel_data_sheet1.append(excel_data_cell1)

    # セル番地を取得 
    for cells in tuple(sheet_set2.rows):
        excel_data_cell2 = []
        for cell in cells:
            excel_data_cell2.append(cell.value)
        excel_data_sheet2.append(excel_data_cell2)        
        
 

    dst_sheet1 = dst_book.create_sheet(index=sheet_number, title=sheet2[sheet_number])
#    
#    dst_sheet1 = dst_book.active
#    dst_sheet1.title = sheet2[sheet_number]
    
    dummy_array = []
    
    max_array_len = max(len(excel_data_sheet1),len(excel_data_sheet2))
    for j in range(len(excel_data_sheet1[sheet_number])) :
        dummy_array.append('dummy')
    
    if (max_array_len - len(excel_data_sheet1)) != 0:
        for i in range(max_array_len - len(excel_data_sheet1)):
            excel_data_sheet1.append(dummy_array)
            
    if (max_array_len - len(excel_data_sheet2)) != 0:
        for i in range(max_array_len - len(excel_data_sheet2)):
            excel_data_sheet2.append(dummy_array)
            
    ###############################################################
    ########新規追加箇所及び変化なし及び部分変更箇所検索
    ###############################################################
    cell_y = 1
#    max_match_array = 0
    
    
    for h in range(max_array_len) :  
        num_match_line1 = 0
        num_match_line2 = 0
        array_match_cell1 = []
        array_match_cell2 = []
        for i in range(max_array_len) :
            num_match_cell1 = 0
            num_match_cell2 = 0
            for j in range(len(excel_data_sheet1[i])) :
            
                if(excel_data_sheet2[i][j] == excel_data_sheet1[h][j]):
                    num_match_cell1 = num_match_cell1 + 1
                if(excel_data_sheet1[i][j] == excel_data_sheet2[h][j]):
                    num_match_cell2 = num_match_cell2 + 1        
                    
            array_match_cell1.append(num_match_cell1)
            array_match_cell2.append(num_match_cell2)
            
            if(num_match_cell1 == len(excel_data_sheet1[i])):
                num_match_line1 = num_match_line1 + 1
            if(num_match_cell2 == len(excel_data_sheet2[i])):
                num_match_line2 = num_match_line2 + 1
            
        ####変化なし箇所
        if num_match_line1 == 1 and num_match_line2 == 1 and excel_data_sheet1[h][0] != 'dummy':
            
    
            cell_x = 65
            print('変化なし行')
            for k in range(len(excel_data_sheet1[h])) :  
                cell_position = chr(cell_x) + str(cell_y)
                dst_sheet1[cell_position]= str(excel_data_sheet1[h][k])
                cell_x = ((cell_x) + 1)
            cell_y = cell_y + 1
                          
            
        ####部分変化箇所
        elif  num_match_line2 == 0 and max(array_match_cell2) >= threshold and excel_data_sheet2[h][0] != 'dummy':
            print('部分変更行')
            
            change_line_number  = array_match_cell2.index(max(array_match_cell2))
            fill = PatternFill(patternType='solid',fgColor='FFFF00', bgColor='FFFF00')
            color_red = Color(rgb='ffff0000')
            font = Font(b=True ,color = color_red)
            cell_x = 65
            for l in range(len(excel_data_sheet1[h])) :  
                
                cell_position = chr(cell_x) + str(cell_y)
                
                if excel_data_sheet2[h][l] != excel_data_sheet1[change_line_number][l]:
                    dst_sheet1[cell_position].font = openpyxl.styles.fonts.Font(color='FF0000')
                    dst_sheet1[cell_position].fill = fill
                    dst_sheet1[cell_position]= str(excel_data_sheet2[h][l])
                    
                else:
                    dst_sheet1[cell_position].fill = fill
                    dst_sheet1[cell_position]= str(excel_data_sheet2[h][l])
                    
                cell_x = ((cell_x) + 1)
            cell_y = cell_y + 1  
            
    
        elif  num_match_line2 == 0 and excel_data_sheet2[h][0] != 'dummy':
            print(str(excel_data_sheet2[h][1]))
            cell_x = 65
            print('新規追加行')
            fill = PatternFill(patternType='solid',fgColor='fd7e00', bgColor='fd7e00')
            for l in range(len(excel_data_sheet2[h])) :  
                cell_position = chr(cell_x) + str(cell_y)
                dst_sheet1[cell_position].fill = fill
                dst_sheet1[cell_position]= str(excel_data_sheet2[h][l])
                cell_x = ((cell_x) + 1)
            cell_y = cell_y + 1
            
        ####削除
        if  num_match_line1 == 0 and  excel_data_sheet1[h][0] != 'dummy':
            cell_x = 65
            print('削除箇所')
            fill = PatternFill(patternType='solid',fgColor='808080', bgColor='808080')
            for l in range(len(excel_data_sheet2[h])) :  
                cell_position = chr(cell_x) + str(cell_y)
                dst_sheet1[cell_position].fill = fill
                dst_sheet1[cell_position]= str(excel_data_sheet1[h][l])
                cell_x = ((cell_x) + 1)
            cell_y = cell_y + 1
            
    
book1.close()
book2.close()    
output_file_name = file_name1[:-5] +("___")+ file_name2[:-5] + ('.xlsx')
dst_book.save(output_file_name)
    
    
